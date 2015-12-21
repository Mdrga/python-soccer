#!python3
# -*- coding: utf-8 -*-
'''
Created on Aug 16, 2015
Modified on Dec 21, 2015
Version 0.03.g
@author: rainier.madruga@gmail.com
A simple Python Program to scrape the ESPN FC website for content.
'''
# Import Libraries needed for Scraping the various web pages
import bs4 
import re
import datetime
import time
import requests
import webbrowser
import os
import openpyxl
import sys
import codecs

# Set Character Output
print ('System Encoding:', sys.stdout.encoding)
sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())

# Establish the process Date & Time Stamp
ts = datetime.datetime.now().strftime("%H:%M:%S")
ds = datetime.datetime.now().strftime("%Y-%m-%d")
date = datetime.datetime.now().strftime("%Y%m%d")

# Updates the Time Stamp
def updateTS():
    update = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return update

# Download Image
def downloadImage(imageURL, localFileName):
    response = requests.get(imageURL)
    if response.status_code == 200:
        print ('Downloading %s...' % (localFileName))
    with open(localimgPath + localFileName, 'wb') as fo:
        for chunk in response.iter_content(4096):
            fo.write(chunk)
    return True

hr = " >>> *** ====================================================== *** <<<"
shr = " >>> *** ==================== *** <<<"

# Program Version & System Variables
parseVersion = 'Premier League Match & Stats Parser v0.03.g'
print (ds + ' :: ' + ts + ' :: ' + parseVersion)
print ('Python Version :: ' + sys.version)
print (hr)

# Define URLs for the Barclay's Premier League
espnURL = 'http://www.espnfc.us/barclays-premier-league/23/index'
espnFixtures = 'http://www.espnfc.us/barclays-premier-league/23/scores?date='
injuriesURL = 'http://www.fantasyfootballscout.co.uk/fantasy-football-injuries/'
teamNewsURL = 'http://www.fantasyfootballscout.co.uk/team-news/'
bbcURL = 'http://www.bbc.com/sport/0/football/premier-league/'
bbcFixturesURL = 'http://www.bbc.com/sport/football/premier-league/fixtures'
bbcResultsURL = 'http://www.bbc.com/sport/football/premier-league/results'

# Base Path for Output
localPath = 'D:\\ESPN-Parser\\'
localimgPath = 'D:\\ESPN-Parser\\imgs\\'
baseWkBk = 'stats_template.xlsx'
workBook = openpyxl.load_workbook(os.path.join(localPath + baseWkBk))
teamSheet = workBook.get_sheet_by_name('teams')
playerSheet = workBook.get_sheet_by_name('players')
matchSheet = workBook.get_sheet_by_name('matches')
fixtureSheet = workBook.get_sheet_by_name('fixtures')

# Create BS4 Object from ESPN Web Page
espnRes = requests.get(espnURL)
espnRes.raise_for_status()
espnSoup = bs4.BeautifulSoup(espnRes.text, "html.parser")
with open(os.path.join(localPath +'espnLanding.txt'), 'wb') as fo:
	for chunk in espnRes.iter_content(100000):
		fo.write(chunk)

espnFixRes = requests.get(espnFixtures)
espnFixRes.raise_for_status()
fixturesSoup = bs4.BeautifulSoup(espnFixRes.text, "html.parser")
with open(os.path.join(localPath + 'fixturesLanding.txt'), 'wb') as fo:
	for chunk in espnFixRes.iter_content(100000):
		fo.write(chunk)

teamListContainer = fixturesSoup.find("div", id="submenu-content")
teamList = teamListContainer.find_all("ul")
teamList = teamList[1].find_all("li")

teamCounter = 2

# Function to return a two digit month for a literal Month (i.e., change "August" to "08").
def returnMonth(x):
	inputMonth = x
	inputMonth = inputMonth[0:3]
	outputMonth = ''
	# print inputMonth
	if inputMonth == 'Aug':
		outputMonth = '08'
	elif inputMonth == 'Sep':
		outputMonth = '09'
	elif inputMonth == 'Oct':
		outputMonth = '10'
	elif inputMonth == 'Nov':
		outputMonth = '11'
	elif inputMonth == 'Dec':
		outputMonth = '12'
	elif inputMonth == 'Jan':
		outputMonth = '01'
	elif inputMonth == 'Feb':
		outputMonth = '02'
	elif inputMonth == 'Mar':
		outputMonth = '03'
	elif inputMonth == 'Apr':
		outputMonth = '04'
	elif inputMonth == 'May':
		outputMonth = '05'
	elif inputMonth == 'Jun':
		outputMonth = '06'
	else:
		outputMonth = '07'
	return outputMonth


# Parse ESPN Team List
for i in teamList:
	teamName = i.get_text()
	teamURL = i.a['href']
	teamSheet.cell('A' + str(teamCounter)).value = teamName
	teamSheet.cell('B' + str(teamCounter)).value = teamURL
	# print (teamName)
	teamCounter += 1

workBook.save(os.path.join(localPath + ds + '.xlsx'))
print ('Teams Saved...')
print (hr)

# Parse BBC Fixtures and Results
bbcFixtures = requests.get(bbcFixturesURL)
bbcFixtures.raise_for_status()

bbcResults = requests.get(bbcResultsURL)
bbcResults.raise_for_status()

bbcFixtureSoup = bs4.BeautifulSoup(bbcFixtures.text, "html.parser")
with open(os.path.join(localPath + 'bbcFixtures.html'), 'wb') as fo:
	for chunk in bbcFixtures.iter_content(100000):
		fo.write(chunk)
		print ('Writing fixtures file...')

print (shr)

bbcResultsSoup = bs4.BeautifulSoup(bbcResults.text, "html.parser")
with open (os.path.join(localPath + 'bbcResults.html'), 'wb') as fo:
	for chunk in bbcResults.iter_content(100000):
		fo.write(chunk)
		print ('Writing results file...')

print (hr)

fixtureContainer = bbcFixtureSoup.find('div', class_="fixtures-table full-table-medium")
fixtureRow = fixtureContainer.find_all('table', class_="table-stats")

resultsContainer = bbcResultsSoup.find('div', class_="fixtures-table full-table-medium")
resultRow = resultsContainer.find_all('table', class_="table-stats")

# Function to process container. Pass two variables. 
# X = Container
# Y = Container Type 1 = Fixtures; 2 = Results
def processContainer(x, y):
	container = x
	processType = y
	count = 2

	if processType == 1:
		print ('Output for Fixtures Started...')
		print (shr)
	if processType == 2:
		print ('Output for Results Started...')
		print (shr)

	for i in container:
		rowData = i
		if len(rowData) == 1:
			fixtureDate = str((rowData.string).strip())
			if len(fixtureDate) > 1:
				dateYear = fixtureDate[-4:]
				dayOfWeek = fixtureDate[0:3]
				dayOfMonth = re.findall(r'\d{1,2}', fixtureDate)
				dayOfMonth = dayOfMonth[0]
				month = re.findall(r'[A-Z][a-z]*', fixtureDate)
				month = month[1]
				printFixtureDate = str(dateYear) + '-' + returnMonth(month[0:3]) + '-' + str(dayOfMonth) # + ' ' + dayOfWeek
				#print (printFixtureDate)
				#print (shr)
		if processType == 1:
			if len(rowData) > 1:
				tableRow = rowData.find_all('tr')
				previewData = i.find_all('tr', class_="preview")
				for i in previewData:
					matchID = i['id']

					# Identify the information about the Home Team
					homeTeam = i.find('span', class_="team-home teams")
					homeTeamURL = homeTeam.a['href']
					homeTeamName = homeTeam.a.get_text()
				
					# Identify the information about the Away Team
					awayTeam = i.find('span', class_="team-away teams")
					awayTeamURL = awayTeam.a['href']
					awayTeamName = awayTeam.a.get_text()
			
					# Identify Kickoff Time and Status
					kickoff = i.find('td', class_="kickoff")
					kickoff = kickoff.get_text().strip()
					status = i.find('td', class_="status")
					status = status.get_text().strip()
					
					# Output to Excel File
					fixtureSheet.cell('A' + str(count)).value = homeTeamName
					fixtureSheet.cell('B' + str(count)).value = awayTeamName
					fixtureSheet.cell('C' + str(count)).value = printFixtureDate
					fixtureSheet.cell('D' + str(count)).value = dayOfWeek
					fixtureSheet.cell('E' + str(count)).value = kickoff
					fixtureSheet.cell('F' + str(count)).value = status
					fixtureSheet.cell('G' + str(count)).value = matchID[14:]

					count += 1

		if processType == 2:
			if len(rowData) > 1:
				tableRow = rowData.find_all('tr', class_="report")
				for i in tableRow:
					# Identify the information about the Home Team
					homeTeam = i.find('span', class_="team-home teams")
					homeTeamURL = homeTeam.a['href']
					homeTeamName = homeTeam.a.get_text()
				
					# Identify the information about the Away Team
					awayTeam = i.find('span', class_="team-away teams")
					awayTeamURL = awayTeam.a['href']
					awayTeamName = awayTeam.a.get_text()

					result = i.find('span', class_="score")
					result = result.get_text()
					result = result.strip()
					if len(result.strip()) <= 3:
						homeScore = result[0]
						awayScore = result[2]
					else:
						scoreLine = result.striip()
						scoreLine.find('-')
						print(len(scoreLine))
						homeScore = '99'
						awayScore = '99'
					matchReport = i.find('a', class_='report')
					matchReport = matchReport['href']
					matchID = matchReport[16:len(matchReport)]

					matchSheet.cell('A' + str(count)).value = homeTeamName
					matchSheet.cell('B' + str(count)).value = awayTeamName
					matchSheet.cell('C' + str(count)).value = printFixtureDate
					matchSheet.cell('D' + str(count)).value = homeScore
					matchSheet.cell('E' + str(count)).value = awayScore
					matchSheet.cell('F' + str(count)).value = "http://www.bbc.com/sport/0/football/" + matchID
					matchSheet.cell('G' + str(count)).value = matchID

					count += 1
	print ('Fixtures and Team Results completed... ')
	workBook.save(os.path.join(localPath + ds + '.xlsx'))
	print (hr)

processContainer(fixtureContainer, 1)
processContainer(resultsContainer, 2)

# Pull URL from Match Sheet and process Player Results
maxResultRow = matchSheet.get_highest_row()

# Parse the match results from a URL for a Side "H"ome or "A"way
def parseResults(url, side):
	matchURL = url
	matchSide = side

playerRow = 2

for row in range(2, matchSheet.get_highest_row()+1):
	homeTeam = matchSheet['A' + str(row)].value
	awayTeam = matchSheet['B' + str(row)].value
	matchURL = matchSheet['F' + str(row)].value
	matchID = matchSheet['G' + str(row)].value
	
	print (homeTeam + ' v. ' + awayTeam)
	print (matchURL)

	# Parse Match Results
	# Have to build in some fault tolerance to this process. Need to be able to parse the 
	# sections and not have the process fall over if the value is not available...
	matchResult = requests.get(matchURL)
	matchResult.raise_for_status()
	matchSoup = bs4.BeautifulSoup(matchResult.text, "html.parser")
	matchUpdate = matchSoup.find('div', id="article-sidebar")
	matchTimestamp = matchUpdate.find('p', class_="page-timestamp")
	matchTimestamp = (matchTimestamp.get_text()).strip()
	matchStats = matchSoup.find('div', id="match-stats-charts")
	matchDetails = matchSoup.find('div', class_="story-body")
	matchLineup = matchSoup.find('div', id="line-up-wrapper")
	
	if matchLineup == None:
		print("Match doesn't have details")
		print(shr)
	else:
		matchReferee = matchLineup.find('div', class_="referee")
		matchReferee = matchReferee.get_text()
		matchReferee = matchReferee[6:len(matchReferee)]
		matchAttendance = matchLineup.find('div', class_="attendance")
		matchAttendance = matchAttendance.get_text()
		matchAttendance = matchAttendance[6:len(matchAttendance)-1]

		# Home Team Lineup
		matchHomeLineup = matchLineup.find('div', class_="home-team")
		matchHomeStarting = matchHomeLineup.find('ul', class_="player-list")
		matchHomeSubs = matchHomeLineup.find('ul', class_="subs-list")
		matchHomeStartingLineup = matchHomeStarting.find_all('li')
		matchHomeSubsLineup = matchHomeSubs.find_all('li')
		maxPlayerRow = playerSheet.get_highest_row()
		if playerRow > 2:
			playerRow = maxPlayerRow
		else:
			playerRow = maxPlayerRow +1

		# Away Team Lineup
		matchAwayLineup = matchLineup.find('div', class_="away-team")
		matchAwayStarting = matchAwayLineup.find('ul', class_="player-list")
		matchAwaySubs = matchHomeLineup.find('ul', class_="subs-list")
		matchAwayStartingLineup = matchAwayStarting.find_all('li')
		matchAwaySubsLineup = matchAwaySubs.find_all('li')

		# Parse Home Team LineUps to Excel Sheet
		for i in matchHomeStartingLineup:
			playerSheet['A' + str(playerRow)].value = matchID
			playerSheet['B' + str(playerRow)].value = matchURL
			playerSheet['C' + str(playerRow)].value = homeTeam
			playerSheet['D' + str(playerRow)].value = 'Home'
			playerSheet['E' + str(playerRow)].value = 'Starter'
			playerSheet['F' + str(playerRow)].value = str(i.get_text())
			print (i.get_text())
			playerRow += 1

		# Parse Away Team Lineup to Excel
		for i in matchAwayStartingLineup:
			playerSheet['A' + str(playerRow)].value = matchID
			playerSheet['B' + str(playerRow)].value = matchURL
			playerSheet['C' + str(playerRow)].value = homeTeam
			playerSheet['D' + str(playerRow)].value = 'Away'
			playerSheet['E' + str(playerRow)].value = 'Starter'
			playerSheet['F' + str(playerRow)].value = str(i.get_text())
			print (i.get_text())
			playerRow += 1

		# Parse Home Team Subs to Excel Sheet
		for i in matchHomeSubsLineup:
			playerSheet['A' + str(playerRow)].value = matchID
			playerSheet['B' + str(playerRow)].value = matchURL
			playerSheet['C' + str(playerRow)].value = homeTeam
			playerSheet['D' + str(playerRow)].value = 'Home'
			playerSheet['E' + str(playerRow)].value = 'Sub'
			playerSheet['F' + str(playerRow)].value = str(i.get_text())
			print (i.get_text())
			playerRow += 1		

		# Parse Away Team Subs to Excel Sheet
		for i in matchAwaySubsLineup:
			playerSheet['A' + str(playerRow)].value = matchID
			playerSheet['B' + str(playerRow)].value = matchURL
			playerSheet['C' + str(playerRow)].value = awayTeam
			playerSheet['D' + str(playerRow)].value = 'Away'
			playerSheet['E' + str(playerRow)].value = 'Sub'
			playerSheet['F' + str(playerRow)].value = str(i.get_text())
			print (i.get_text())
			playerRow += 1		


		print ('>>>======<<<')
		matchHomeTeamScorers = []

		# Update Match Sheet
		matchSheet.cell('H' + str(row)).value = matchReferee
		matchSheet.cell('I' + str(row)).value = matchAttendance

	# Home Team Details
	matchHomeTeam = matchDetails.find('div', id="home-team")
	matchHomeTeamBadge = matchHomeTeam.find('div', class_="team-badge")
	matchHomeTeamBadge = matchHomeTeamBadge.find('img')
	matchHomeTeamBadge = matchHomeTeamBadge['src']
	matchHomeTeamScore = matchHomeTeam.find('span', class_="team-score")
	homeTeamScore = (matchHomeTeamScore.get_text()).strip()
	matchHomeTeamSpans = matchHomeTeam.find_all('span')
		
	print ('>>>======<<<')
	# print (matchID)
	
	# print (shr)
	# print (matchHomeSubs.prettify())

	# Away Team Details
	matchAwayTeam = matchDetails.find('div', id="away-team")
	matchAwayTeamBadge = matchAwayTeam.find('div', class_="team-badge")
	matchAwayTeamBadge = matchAwayTeamBadge.find('img')
	matchAwayTeamBadge = matchAwayTeamBadge['src']

	# Download Home Team Badge
	if (os.path.isfile(localimgPath + homeTeam + '.png')) == False:
		downloadImage(matchHomeTeamBadge, homeTeam + '.png')
	if (os.path.isfile(localimgPath + awayTeam + '.png')) == False:		
		downloadImage(matchAwayTeamBadge, awayTeam + '.png')

	#print (matchTimestamp)
	# print (matchLineup.prettify())
	# print (matchDetails.prettify())
	# print (' >>>***========***<<< ')
	counter = 2
	if len(matchHomeTeamSpans) >= 3:
		maxLen = len(matchHomeTeamSpans)
		for i in matchHomeTeamSpans[counter:maxLen]:
			goalScorer =  i.get_text().strip()
			matchHomeTeamScorers.append(goalScorer)
			# print (counter)
			counter += 1
	print (matchHomeTeamScorers)
	# print (' >>>***========***<<< ')
	# print (len(matchHomeTeamSpans))
	# print (matchHomeTeamSpans)
	print (' >>>***========***<<< ')
	#print (matchHomeTeamBadge)
	#print (' >>>***========***<<< ')

	
	'''
	# Find home team, away team and team badges:
	with open (os.path.join(localPath + 'data\\' + matchID + '.html'), 'wb') as fo:
		for chunk in matchResult.iter_content(100000):
			fo.write(chunk)
			print ('Writing match # '+ matchID + ' results file...')
	'''
	print (shr)
	workBook.save(os.path.join(localPath + ds + '.xlsx'))


workBook.save(os.path.join(localPath + ds + '.xlsx'))
print ('Max Result Row is: ' + str(maxResultRow))

print (hr)